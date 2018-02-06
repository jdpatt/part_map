''' Visual pinout of a bga '''
from tkinter import Tk, Canvas, Scrollbar, RIGHT, Y
from re import split
from logging import basicConfig, getLogger, INFO
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from natsort import natsorted
from argparse import ArgumentParser


def sort_and_split_pin_list(plist):
    ''' Haven't found a cleaner way to sort. '''
    r_list = list()
    c_list = list()
    for pin in plist:
        split_pin = split(r'(\d+)', pin)
        if split_pin[0] not in r_list:
            r_list.append(split_pin[0])
        c_list.append(split_pin[1])
    temp = list()
    temp2 = list()
    for item in r_list:
        if len(item) > 1:
            temp.append(item)
        else:
            temp2.append(item)
    temp2 = natsorted(temp2)
    temp2.extend(natsorted(temp))
    return natsorted(set(c_list)), temp2


def read_in_xlsx_file(file):
    '''  Read in excel and get the pin number, function, and color '''
    workbook = load_workbook(file)
    sheet = workbook.active  # Grab the first sheet
    for col in sheet.iter_cols():
        for cell in col:
            if cell.value == 'Name':
                function_column = column_index_from_string(cell.column)
            if cell.value == 'Number':
                pin_column = column_index_from_string(cell.column)
    bga_dict = dict()
    pin_list = list()
    color_map_dict = dict()
    for excel_row in range(2, sheet.max_row + 1):
        pin = sheet.cell(row=excel_row, column=pin_column)
        function = sheet.cell(row=excel_row, column=function_column)
        if pin.value is not None or function.value is not None:
            bga_dict.update({pin.value: function.value})
            pin_list.append(pin.value)
            if function.value not in color_map_dict:
                if function.fill.patternType == 'solid':
                    color_map_dict.update({pin.value: '#' + function.fill.start_color.rgb[2:]})
    col_list, row_list = sort_and_split_pin_list(pin_list)
    return bga_dict, col_list, row_list, color_map_dict


if __name__ == '__main__':
    GRID_SPACING = 0
    ELEMENT_SIZE = 24
    BORDER = 30
    OFFSET = BORDER + GRID_SPACING

    PARSER = ArgumentParser()
    PARSER.add_argument('excel_file',
                        nargs='?',
                        default='artix7_example.xlsx',
                        help='The excel file to read in')
    PARSER.add_argument('--circles', '-c',
                        action='store_true',
                        help='Print circles instead of rectangles')
    ARGS = PARSER.parse_args()
    BGA, ROWS, COLUMNS, COLORS = read_in_xlsx_file(ARGS.excel_file)

    MASTER = Tk()
    PIN_SCROLL = Scrollbar(MASTER)
    PIN_SCROLL.pack(side=RIGHT, fill=Y)
    WINDOW = Canvas(MASTER, width=((len(COLUMNS) + 1) * ELEMENT_SIZE) + BORDER,
                    height=(len(ROWS) + 1) * ELEMENT_SIZE + BORDER)
    WINDOW.pack()
    PIN_SCROLL.config(command=WINDOW.yview)

    for hdr_index, column in enumerate(COLUMNS):
        WINDOW.create_text((ELEMENT_SIZE * hdr_index) + OFFSET +
                           GRID_SPACING + (ELEMENT_SIZE/2),
                           GRID_SPACING + (ELEMENT_SIZE/2),
                           font=("Purisa", 4),
                           text=column)

    try:
        # Create the BGA Canvas
        for r_index, row in enumerate(ROWS):
            for c_index, column in enumerate(COLUMNS):
                pinname = column + str(row)
                if pinname in BGA:
                    if BGA[pinname] is not None and len(BGA[pinname]) > 5:
                        pin_function = BGA[pinname][:5]
                    else:
                        pin_function = BGA[pinname]
                    if pinname in COLORS:
                        fill = COLORS[pinname]
                    else:
                        fill = ''
                    if ARGS.circles:
                        WINDOW.create_oval((ELEMENT_SIZE * c_index) + OFFSET,
                                           (ELEMENT_SIZE * r_index) + OFFSET,
                                           (ELEMENT_SIZE * c_index) + ELEMENT_SIZE + OFFSET,
                                           (ELEMENT_SIZE * r_index) + ELEMENT_SIZE + BORDER,
                                           fill=fill)
                    else:
                        WINDOW.create_rectangle((ELEMENT_SIZE * c_index) + OFFSET,
                                                (ELEMENT_SIZE * r_index) + OFFSET,
                                                (ELEMENT_SIZE * c_index) + ELEMENT_SIZE + OFFSET,
                                                (ELEMENT_SIZE * r_index) + ELEMENT_SIZE + BORDER,
                                                fill=fill)
                    WINDOW.create_text((ELEMENT_SIZE * c_index) + OFFSET +
                                       GRID_SPACING + (ELEMENT_SIZE/2),
                                       (ELEMENT_SIZE * r_index) + OFFSET + (ELEMENT_SIZE/2),
                                       font=("Purisa", 4),
                                       text=pin_function)
            WINDOW.create_text((ELEMENT_SIZE * (len(COLUMNS))) + OFFSET + (ELEMENT_SIZE/2),
                               (ELEMENT_SIZE * r_index) + OFFSET + (ELEMENT_SIZE/2),
                               font=("Purisa", 4),
                               text=row)
        MASTER.mainloop()
    except ValueError:
        basicConfig(filename="bga_visual.log",
                    filemode="w",
                    level=INFO)
        LOG = getLogger()
        LOG.error(BGA)
        LOG.error(ROWS)
        LOG.error(COLUMNS)
        LOG.error(COLORS)
