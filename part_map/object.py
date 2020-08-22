"""Class representing the object being modeled."""
import json
import logging
import re
from pathlib import Path
from typing import Dict, List, Tuple, Union

from natsort import natsorted
from openpyxl import load_workbook


class PartObject:
    """ Load and create a part from a source """

    def __init__(self, pins, filename):
        super().__init__()
        self.log = logging.getLogger("partmap.object")
        self._pins = pins
        self._columns, self._rows = self.sort_and_split_pin_list()
        self.filename = Path(filename)

    @classmethod
    def from_excel(cls, filename):
        """ Import an Excel and create a PartObject """
        number = "Number"
        name = "Name"
        workbook = load_workbook(filename)
        sheet = workbook.active  # Grab the first sheet
        try:
            column = get_col_index([number, name], sheet)
            bga = dict()
            for excel_row in range(2, sheet.max_row + 1):
                pin = sheet.cell(row=excel_row, column=column[number]).value
                net = sheet.cell(row=excel_row, column=column[name])
                if pin is not None or net.value is not None:
                    if net.fill.patternType == "solid":
                        bga.update(
                            {
                                pin: {
                                    "name": net.value,
                                    "color": str("#" + net.fill.start_color.rgb[2:]),
                                }
                            }
                        )
                    else:
                        bga.update({pin: {"name": net.value, "color": "#ffffff"}})
        except (TypeError, ValueError, KeyError, UnboundLocalError) as error:
            print(error)
            raise
        return cls(bga, filename)

    @classmethod
    def from_telesis(cls, filename, refdes):
        """ Import a Telesis formatted file and create a PartObject """
        with open(filename, "r") as tel_file:
            tel_text = tel_file.readlines()
        tel_netlist = dict()
        for line in tel_text:
            reg = re.match(r"(.*);", line)
            reg2 = re.findall(refdes + r"\.([a-zA-Z0-9]+)", line)
            if reg and reg2:
                net = reg.group(1)
                for reg_match in reg2:
                    pin = reg_match
                    tel_netlist.update({pin: {"name": net, "color": "#ffffff"}})
        return cls(tel_netlist, filename)

    @classmethod
    def from_json(cls, filename):
        """ Import a json file with a format {pin: {name:, color:}} """
        return cls(json.load(open(filename)), filename)

    def add_pin(self, pin: str, net: str, color: str) -> None:
        """ Add a new pin to the part.

        Args:
            pin: The Pin Number. (A12)
            net: The functional name of the net. (USB_P)
            color: The color to fill with.
        """
        self._pins.update({pin: {"name": net, "color": color}})

    @property
    def columns(self) -> List:
        """ Get the columns in a part. [1-n] """
        return self._columns

    @columns.setter
    def columns(self, new_columns):
        """Update the columns."""
        self._columns = new_columns

    @property
    def rows(self) -> List:
        """ Get the rows in a part.  [A - AZ] """
        return self._rows

    @rows.setter
    def rows(self, new_rows):
        """Update the rows."""
        self._rows = new_rows

    def get_pin(self, prefix: str, suffix: str) -> Union[str, None]:
        """ Get the name and color of a pin """
        pin = None
        if prefix + suffix in self._pins:
            pin = self._pins[prefix + suffix]
        elif suffix + prefix in self._pins:
            pin = self._pins[suffix + prefix]
        return pin

    @property
    def pins(self):
        """ Return the pin names """
        return self._pins.keys()

    def get_number_of_pins(self):
        """ Return how many pins are in the part """
        return len(self._pins)

    def get_net_names(self):
        """ Return the net names """
        return self._pins.values()["name"]

    def dump_json(self):
        """ Dump the PartObject dictionary to a .json file """
        save_file = self.filename.with_suffix(".json")
        self.log.info(f"Saved as json to {save_file}")
        with open(save_file, "w") as outfile:
            json.dump(self._pins, outfile, sort_keys=True, indent=4, separators=(",", ": "))

    def sort_and_split_pin_list(self) -> Tuple[List, List]:
        """ Take a list of pins and spilt by letter and number then sort """
        r_list: List = list()
        c_list = list()
        for pin in self.pins:
            split_pin = re.split(r"(\d+)", pin)
            if split_pin[0] not in r_list:
                r_list.append(split_pin[0])
            c_list.append(split_pin[1])
        temp = list()
        temp2 = list()
        for item in r_list:
            if len(item) > 1:
                temp.append(item)
            else:
                temp2.append(item)
        temp2 = natsorted(temp2)
        temp2.extend(natsorted(temp))
        return natsorted(set(c_list)), temp2


def get_col_index(name: List, worksheet) -> Dict:
    """ return a list of the column numbers if it matches """
    indexes = dict()
    for rows in worksheet.iter_rows(min_row=1, max_row=1, min_col=1):
        for column in rows:
            if column.value in name:
                indexes.update({column.value: column.col_idx})
    return indexes
