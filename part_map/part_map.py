""" Visual pin out of a BGA or connector """
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Union

import click

from natsort import natsorted
from openpyxl import load_workbook
from PySide2.QtCore import QRectF, Qt
from PySide2.QtGui import QBrush, QColor, QImage, QKeySequence, QPainter, QPixmap
from PySide2.QtWidgets import QAction, QApplication, QGraphicsScene, QGraphicsView, QMenu


class PartViewer(QGraphicsView):
    """ Create a render of the part and load it into a QWidget """

    def __init__(self, part, partview_settings):
        super(PartViewer, self).__init__()
        self.settings = partview_settings
        self.part = part
        self.image = None
        self.box_size = 50
        self.total_steps = 0
        self.generate_render()

    def initUI(self) -> None:  # pylint: disable=C0103
        """ Init all the UI elements """
        self.setWindowTitle(self.settings["title"])

        self.save_action = QAction(
            "Save as Png", self, shortcut=QKeySequence.Save, statusTip="Save to .png", triggered=self.save
        )

        self.toggle_action = QAction(
            "Toggle Shape",
            self,
            shortcut="Ctrl+R",
            statusTip="Switch styles between circles or squares and redraw.",
            triggered=self.toggle_style,
        )

        self.rotate_action = QAction(
            "Rotate",
            self,
            shortcut="Ctrl+R",
            statusTip="Rotate the diagram 90 degrees and redraw.",
            triggered=self.rotate_drawing,
        )

        self.dump_action = QAction(
            "Save as Json",
            self,
            shortcut="Ctrl+Shift+S",
            statusTip="Save the partObject as a json file.",
            triggered=self.part.dump_json,
        )

        self.resize(self.settings["width"], self.settings["height"])
        self.scene = QGraphicsScene()
        self.redraw()
        self.setScene(self.scene)
        self.setTransformationAnchor(self.AnchorUnderMouse)
        self.setResizeAnchor(self.AnchorUnderMouse)
        self.setDragMode(self.ScrollHandDrag)
        self.fitInView(
            QRectF(0, 0, self.settings["width"], self.settings["height"]), Qt.KeepAspectRatio
        )
        self.show()

    def generate_render(self) -> None:
        """ Generate the part """
        part_cols = self.part.get_columns()
        part_rows = self.part.get_rows()
        self.scale_box_size(part_cols, part_rows)
        self.image = QImage(
            self.settings["image_width"],
            self.settings["image_height"] + (self.box_size / 2),
            QImage.Format_RGB32,
        )
        self.image.fill(QColor("#ffffff"))  # Background Color
        paint = QPainter(self.image)
        paint.setRenderHints(
            QPainter.HighQualityAntialiasing
            | QPainter.SmoothPixmapTransform
            | QPainter.TextAntialiasing,
            True,
        )
        if self.settings["rotate"]:
            part_cols.reverse()
            part_cols, part_rows = part_rows, part_cols
        for hdr_offset, column in enumerate(part_cols):
            paint.drawText(
                QRectF(
                    hdr_offset * self.box_size + int(self.box_size / 2),
                    0,
                    self.box_size,
                    self.box_size,
                ),
                Qt.AlignCenter,
                column,
            )
        for y_offset, row in enumerate(part_rows):
            for x_offset, column in enumerate(part_cols):
                pin = self.part.get_pin(str(row), str(column))
                if pin and pin["name"] is not None:
                    fill = pin["color"]
                    brush = QBrush(QColor(fill))
                    paint.setBrush(brush)
                    rect = QRectF(
                        self.box_size * x_offset + int(self.box_size / 2),
                        self.box_size * y_offset + self.box_size,
                        self.box_size,
                        self.box_size,
                    )
                    if self.settings["circles"]:
                        paint.drawEllipse(rect)
                    else:
                        paint.drawRect(rect)
                    if not self.settings["labels"]:
                        paint.drawText(rect, Qt.AlignCenter, pin["name"][:6])
            paint.drawText(
                QRectF(
                    self.box_size * len(part_cols) + int(self.box_size / 2),
                    self.box_size * y_offset + self.box_size,
                    self.box_size,
                    self.box_size,
                ),
                Qt.AlignCenter,
                row,
            )
        paint.end()

    def save(self) -> None:
        """ Save the Pixmap as a .png """
        save_file = PATH.joinpath(f'{self.settings["title"]}.png')
        print(f"Saved to {save_file}")
        self.image.save(str(save_file))

    def scale_box_size(self, columns: List, rows: List) -> None:
        """ If the part width is less than 1536 (2K width) scale up
        """
        part_width = (len(columns) + 1) * self.box_size
        if part_width < 1536.0:
            window_scale = 1536.0 / part_width
        else:
            window_scale = 1
        self.box_size = int(self.box_size * window_scale)
        self.settings["image_width"] = (len(columns) + 1) * self.box_size + self.box_size
        self.settings["image_height"] = (len(rows) + 1) * self.box_size + self.box_size

    def redraw(self):
        """Update the current pixmap."""
        self.scene.clear()
        pixmap = QPixmap.fromImage(self.image)
        self.scene.addPixmap(pixmap)

    def toggle_style(self):
        """Change between circles or squares."""
        if self.settings["circles"]:  # If true, set to false for rectangles.
            self.settings["circles"] = False
        else:
            self.settings["circles"] = True
        self.generate_render()
        self.redraw()

    def rotate_drawing(self):
        """Rotate the diagram."""
        if self.settings["rotate"]:
            self.settings["rotate"] = False
        else:
            self.settings["rotate"] = True
        self.generate_render()
        self.redraw()

    def contextMenuEvent(self, event):
        """ Add the ability to expose the command line switches via a context menu."""
        menu = QMenu(self)
        menu.addAction(self.save_action)
        menu.addAction(self.dump_action)
        menu.addSeparator()
        menu.addAction(self.toggle_action)
        menu.addAction(self.rotate_action)
        menu.exec_(event.globalPos())

    def wheelEvent(self, event) -> None:  # pylint: disable=C0103
        """ If the wheel is scrolled; figure out how much to zoom """
        steps = event.angleDelta().y() / 120
        self.total_steps += steps
        if (self.total_steps * steps) < 0:
            self.total_steps = steps  # Zoom out by steps
        factor = 1.0 + (steps / 10.0)
        self.settings["factor"] *= factor
        if self.settings["factor"] < 0.3:
            self.settings["factor"] = 0.3
        elif self.settings["factor"] > 5:
            self.settings["factor"] = 5
            return
        else:
            self.scale(factor, factor)


class PartObject:
    """ Load and create a part from a source """

    def __init__(self, pins, filename):
        super(PartObject, self).__init__()
        self.__pins = pins
        self.__columns, self.__rows = self.sort_and_split_pin_list()
        self.filename = Path(filename).stem

    @classmethod
    def from_excel(cls, filename):
        """ Import an Excel and create a PartObject """
        number = "Number"
        name = "Name"
        workbook = load_workbook(filename)
        sheet = workbook.active  # Grab the first sheet
        try:
            column = get_col_index([number, name], sheet)
            bga = dict()
            for excel_row in range(2, sheet.max_row + 1):
                pin = sheet.cell(row=excel_row, column=column[number]).value
                net = sheet.cell(row=excel_row, column=column[name])
                if pin is not None or net.value is not None:
                    if net.fill.patternType == "solid":
                        bga.update(
                            {
                                pin: {
                                    "name": net.value,
                                    "color": str("#" + net.fill.start_color.rgb[2:]),
                                }
                            }
                        )
                    else:
                        bga.update({pin: {"name": net.value, "color": "#ffffff"}})
        except (TypeError, ValueError, KeyError, UnboundLocalError) as error:
            print(error)
            raise
        return cls(bga, filename)

    @classmethod
    def from_telesis(cls, filename, refdes):
        """ Import a Telesis formatted file and create a PartObject """
        with open(filename, "r") as tel_file:
            tel_text = tel_file.readlines()
        tel_netlist = dict()
        for line in tel_text:
            reg = re.match(r"(.*);", line)
            reg2 = re.findall(refdes + r"\.([a-zA-Z0-9]+)", line)
            if reg and reg2:
                net = reg.group(1)
                for reg_match in reg2:
                    pin = reg_match
                    tel_netlist.update({pin: {"name": net, "color": "#ffffff"}})
        return cls(tel_netlist, filename)

    @classmethod
    def from_json(cls, filename):
        """ Import a json file with a format {pin: {name:, color:}} """
        return cls(json.load(open(filename)), filename)

    def add_pin(self, pin: str, net: str, color: str) -> None:
        """ Add a new pin to the part.
        
        Args:
            pin: The Pin Number. (A12)
            net: The functional name of the net. (USB_P)
            color: The color to fill with.
        """
        self.__pins.update({pin: {"name": net, "color": color}})

    def get_columns(self) -> List:
        """ Get the columns in a part. [1-n] """
        return self.__columns

    def get_rows(self) -> List:
        """ Get the rows in a part.  [A - AZ] """
        return self.__rows

    def get_pin(self, prefix: str, suffix: str) -> Union[str, None]:
        """ Get the name and color of a pin """
        pin = None
        if prefix + suffix in self.__pins:
            pin = self.__pins[prefix + suffix]
        elif suffix + prefix in self.__pins:
            pin = self.__pins[suffix + prefix]
        return pin

    def get_pins(self):
        """ Return the pin names """
        return self.__pins.keys()

    def get_number_of_pins(self):
        """ Return how many pins are in the part """
        return len(self.__pins)

    def get_net_names(self):
        """ Return the net names """
        return self.__pins.values()["name"]

    def dump_json(self):
        """ Dump the PartObject dictionary to a .json file """
        save_file = PATH.joinpath(f"{self.filename}.json")
        print(f"Saved to {save_file}")
        with open(save_file, "w") as outfile:
            json.dump(self.__pins, outfile, indent=4, separators=(",", ": "))

    def sort_and_split_pin_list(self) -> Tuple[List, List]:
        """ Take a list of pins and spilt by letter and number then sort """
        r_list: List = list()
        c_list = list()
        for pin in self.get_pins():
            split_pin = re.split(r"(\d+)", pin)
            if split_pin[0] not in r_list:
                r_list.append(split_pin[0])
            c_list.append(split_pin[1])
        temp = list()
        temp2 = list()
        for item in r_list:
            if len(item) > 1:
                temp.append(item)
            else:
                temp2.append(item)
        temp2 = natsorted(temp2)
        temp2.extend(natsorted(temp))
        return natsorted(set(c_list)), temp2


def get_col_index(name: List, worksheet) -> Dict:
    """ return a list of the column numbers if it matches """
    indexes = dict()
    for rows in worksheet.iter_rows(min_row=1, max_row=1, min_col=1):
        for column in rows:
            if column.value in name:
                indexes.update({column.value: column.col_idx})
    return indexes


@click.command(context_settings=dict(help_option_names=["-h", "--help"]))
@click.version_option()
@click.argument("filename", type=click.Path(exists=True))
@click.argument("file-type", type=click.Choice(["json", "tel", "excel"]))
@click.option("--refdes", help="The refdes to pull from the Telesis.")
@click.option("--circles", "-c", is_flag=True, help="Draw using circles instead of rectangles.")
# @click.option("--crosshair", is_flag=True, help="Divide the diagram into four quadrants.")
@click.option("--rotate", "-r", is_flag=True, help="Rotate the image by 90 degrees.")
@click.option("--no-labels", is_flag=True, help="Disable the text labels.")
@click.option("--save", "-s", is_flag=True, help="Save the image as a .png.")
@click.option("--dump", "-d", is_flag=True, help="Dump PartObject as a Json File.")
@click.option("--nogui", "-n", is_flag=True, help="Do not open GUI window.")
def cli(**kwargs) -> None:
    """Generate a visualization of a part to help with pinout or general planning.
    Helpful in the early stages of design before layout begins.  Can use it for
    anything that is on a grid.
    """
    print(kwargs)
    filename = Path(kwargs["filename"])
    if kwargs["file_type"] == "excel":
        part = PartObject.from_excel(filename)
    elif kwargs["file_type"] == "json":
        part = PartObject.from_json(filename)
    elif kwargs["file_type"] == "tel":
        part = PartObject.from_telesis(filename, kwargs["refdes"])

    global APP  # pylint: disable=w0603
    APP = QApplication(sys.argv)
    screen_resolution = APP.desktop().screenGeometry()
    view_settings = {
        "width": screen_resolution.width(),
        "height": screen_resolution.height(),
        "title": filename.stem,
        "rotate": kwargs["rotate"],
        "circles": kwargs["circles"],
        # "crosshair": kwargs["crosshair"],
        "labels": kwargs["no_labels"],
        "factor": 1.0,
    }

    view = PartViewer(part, view_settings)

    if not kwargs["nogui"]:
        view.initUI()
    if kwargs["dump"]:
        part.dump_json()
    if kwargs["save"]:
        view.save()
    if kwargs["nogui"]:
        APP.closeAllWindows()
    else:
        sys.exit(APP.exec_())


APP = None
PATH = Path.cwd()


if __name__ == "__main__":
    cli()
